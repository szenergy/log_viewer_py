#!/usr/bin/env python3
"""Verify the generated Excel file."""

import openpyxl

# Open the generated Excel file
wb = openpyxl.load_workbook('output/Test_17_06_2026_05_46_54.xlsx')

print("Workbook sheets:", wb.sheetnames)
print()

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"Sheet: {sheet_name}")
    print(f"  Dimensions: {ws.dimensions}")
    print(f"  Rows: {ws.max_row}, Columns: {ws.max_column}")
    headers = [cell.value for cell in ws[1]]
    print(f"  Columns ({len(headers)}): {', '.join(headers[:5])}... (and {len(headers)-5} more)")
