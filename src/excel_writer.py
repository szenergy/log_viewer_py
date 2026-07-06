"""Excel writer module for converting TDMS data to Excel format."""

import sys
from typing import Dict, List
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows


def print_progress(current: int, total: int, prefix: str = "", suffix: str = "") -> None:
    """
    Print a progress bar to stdout.
    
    Args:
        current: Current progress value
        total: Total progress value
        prefix: Prefix text
        suffix: Suffix text
    """
    if total == 0:
        return
    
    percent = current / total
    bar_length = 20
    filled = int(bar_length * percent)
    bar = '█' * filled + '░' * (bar_length - filled)
    
    status = f"{percent*100:.0f}%"
    output = f"\r{prefix} [{bar}] {status} {suffix}"
    
    # Use sys.stdout.write and flush for better control
    sys.stdout.write(output)
    sys.stdout.flush()
    
    # Print newline when complete
    if current >= total:
        print()


def create_workbook() -> Workbook:
    """
    Create a new Excel workbook.
    
    Returns:
        New Workbook object
    """
    workbook = Workbook()
    # Remove default sheet
    if 'Sheet' in workbook.sheetnames:
        del workbook['Sheet']
    return workbook


def add_data_sheet(workbook: Workbook, sheet_name: str, data: pd.DataFrame, show_progress: bool = False) -> None:
    """
    Add a sheet with data to the workbook.
    
    Args:
        workbook: Workbook object
        sheet_name: Name for the new sheet
        data: DataFrame with data to write
        show_progress: Show progress indicator while writing
    """
    # Create sheet
    worksheet = workbook.create_sheet(sheet_name)
    
    # Write DataFrame to sheet with headers
    rows = list(dataframe_to_rows(data, index=False, header=True))
    total_rows = len(rows)
    
    for r_idx, row in enumerate(rows, 1):
        if show_progress and r_idx % max(1, total_rows // 100) == 0:
            print_progress(r_idx, total_rows, prefix="    Writing rows...")
        
        for c_idx, value in enumerate(row, 1):
            worksheet.cell(row=r_idx, column=c_idx, value=value)
    
    if show_progress:
        print_progress(total_rows, total_rows, prefix="    Writing rows...")
    
    # Auto-adjust column widths
    for column in worksheet.columns:
        max_length = 0
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)  # Cap at 50
        worksheet.column_dimensions[column[0].column_letter].width = adjusted_width


def save_workbook(workbook: Workbook, filepath: str) -> None:
    """
    Save workbook to file.
    
    Args:
        workbook: Workbook object
        filepath: Path where to save the Excel file
    """
    workbook.save(filepath)
